"""Extract relay URLs from a nostr.watch xlsx export.

    python -m src.nostr.extract_relays <xlsx> <out.txt> [--online] [--clearnet]
      --online    keep only in_rstate=True (active rotation)
      --clearnet  drop tor/i2p

Requires `openpyxl` (an optional dependency — only needed for this helper).
Ported from `../nostr-cf-recon/extract_relays.py`.
"""
from __future__ import annotations

import sys
from typing import List


def extract(src: str, dst: str, only_online: bool = False, only_clearnet: bool = False) -> int:
    from openpyxl import load_workbook  # imported lazily so the dep stays optional

    wb = load_workbook(src, read_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    header = list(next(rows))
    url_idx = header.index("url")
    # Only the requested filters need their columns — look them up lazily so a
    # no-flag run doesn't crash on an export that omits in_rstate/network.
    rstate_idx = header.index("in_rstate") if only_online else None
    net_idx = header.index("network") if only_clearnet else None

    seen = set()
    with open(dst, "w") as f:
        for row in rows:
            url = row[url_idx]
            if not url:
                continue
            if only_online and not row[rstate_idx]:
                continue
            if only_clearnet and row[net_idx] != "clearnet":
                continue
            url = str(url).strip()
            if url in seen:
                continue
            seen.add(url)
            f.write(url + "\n")
    return len(seen)


def main(argv: List[str] | None = None) -> int:
    args = sys.argv[1:] if argv is None else argv
    flags = {a for a in args if a.startswith("--")}
    pos = [a for a in args if not a.startswith("--")]
    if len(pos) < 2:
        print("usage: extract_relays.py <xlsx> <out.txt> [--online] [--clearnet]")
        return 1
    src, dst = pos[0], pos[1]
    count = extract(src, dst, only_online="--online" in flags, only_clearnet="--clearnet" in flags)
    print(f"wrote {count} urls -> {dst}  (online={'--online' in flags}, clearnet={'--clearnet' in flags})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
